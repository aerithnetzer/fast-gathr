from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tagger.models import StageOutput

from .contracts import ExecutionStatus
from .entity_review_handoff import build_entity_downstream_package, entity_downstream_is_eligible
from .providers.factory import StageGenerationClient, stage_generation_client
from .stage_runner import _document_header_and_body


PROJECT_ROOT = Path(__file__).resolve().parents[2]
CHATBOT_DIR = PROJECT_ROOT.parent / "Chatbot docs" / "Claude chatbots"
LEGACY_KEYWORD_REGISTRY = PROJECT_ROOT / "resources" / "legacy" / "Keyword_Registry_old.xlsx"
SUMMARY_PROMPT = CHATBOT_DIR / "Keyword_SystemPrompt_v3_KB.txt"
SUMMARY_GUIDE = CHATBOT_DIR / "Keyword_Reference_Guide.txt"
ASSEMBLER_PROMPT = CHATBOT_DIR / "Assembler_System_Prompt.txt"
ASSEMBLER_INSTRUCTIONS = CHATBOT_DIR / "Assembly_Instructions.txt"

SUMMARY_CONTRACT = "summary-keywords-legacy-registry-v1"
ASSEMBLER_CONTRACT = "tag-assembler-occurrence-conservation-v1"
TAG_LINE = re.compile(r"^\s*[EAQ]:\s*.+$", re.MULTILINE)


class SummaryAssemblerError(RuntimeError):
    pass


@dataclass(frozen=True)
class GenerationResult:
    status: str
    raw_output: str
    payload: dict[str, Any]
    provenance: dict[str, Any]
    validation: dict[str, Any]
    provider: str
    model: str
    error: str = ""
    request: dict[str, Any] | None = None


def load_legacy_keyword_registry(path: Path = LEGACY_KEYWORD_REGISTRY) -> dict[str, Any]:
    if not path.exists():
        raise SummaryAssemblerError(f"Legacy Keyword Registry is missing: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    records = [
        {headers[index]: str(value or "").strip() for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "authority_status": "legacy_temporary_authority",
        "filename": path.name,
        "sha256": digest,
        "headers": headers,
        "records": records,
    }


def validate_summary_keywords_output(
    raw_output: str, *, source_body: str, registry: dict[str, Any]
) -> dict[str, Any]:
    text = str(raw_output or "").strip()
    summary_match = re.search(r"(?is)^\s*SUMMARY\s*\n+(.*?)\n+KEYWORDS\s*\n+(.*)$", text)
    issues: list[dict[str, str]] = []
    if not summary_match:
        return {"valid": False, "issues": [{"code": "sections_missing"}], "keywords": []}
    summary = summary_match.group(1).strip()
    keyword_text = summary_match.group(2).strip()
    known_ids = {str(row.get("ID") or "") for row in registry["records"]}
    keyword_ids = re.findall(r"\bK-\d{4}\b", keyword_text)
    unknown_ids = sorted(set(keyword_ids) - known_ids)
    if unknown_ids:
        issues.append({"code": "unknown_legacy_keyword_ids", "detail": ", ".join(unknown_ids)})
    evidence_values = re.findall(r"(?im)^\s*Evidence:\s*[\"“]?(.+?)[\"”]?\s*$", keyword_text)
    missing_evidence = []
    for evidence in evidence_values:
        fragments = [part.strip(" \t\"“”") for part in re.split(r"\s*(?:\.\.\.|…)\s*", evidence)]
        if any(fragment and fragment not in source_body for fragment in fragments):
            missing_evidence.append(evidence)
    if missing_evidence:
        issues.append({"code": "evidence_not_verbatim", "detail": str(len(missing_evidence))})
    if not summary:
        issues.append({"code": "summary_empty"})
    if not keyword_text:
        issues.append({"code": "keywords_empty"})
    return {
        "valid": not issues,
        "issues": issues,
        "summary": summary,
        "summary_word_count": len(summary.split()),
        "keyword_ids": keyword_ids,
        "evidence_count": len(evidence_values),
    }


def validate_occurrence_conservation(occurrence_input: str, assembler_output: str) -> dict[str, Any]:
    input_lines = [line.strip() for line in TAG_LINE.findall(str(occurrence_input or ""))]
    output_lines = [line.strip() for line in TAG_LINE.findall(str(assembler_output or ""))]
    missing = [line for line in input_lines if output_lines.count(line) < input_lines.count(line)]
    added = [line for line in output_lines if input_lines.count(line) < output_lines.count(line)]
    preserved_order = [line for line in output_lines if line in input_lines] == input_lines
    return {
        "valid": bool(input_lines) and not missing and not added and preserved_order,
        "input_tag_count": len(input_lines),
        "output_tag_count": len(output_lines),
        "missing_or_modified_lines": sorted(set(missing)),
        "added_eaq_lines": sorted(set(added)),
        "input_order_preserved": preserved_order,
        "comparison": "exact stripped E/A/Q line equality",
    }


class SummaryKeywordsGenerationService:
    def __init__(self, *, client: StageGenerationClient | None = None) -> None:
        self.client = client or stage_generation_client()

    def run(self, *, document: Any, request_id: str, max_output_tokens: int = 768) -> GenerationResult:
        registry = load_legacy_keyword_registry()
        header, source_body = _document_header_and_body(document)
        system_prompt = SUMMARY_PROMPT.read_text(encoding="utf-8-sig")
        guide = SUMMARY_GUIDE.read_text(encoding="utf-8-sig")
        headers = registry["headers"]
        registry_text = "\n".join(
            " | ".join(str(row.get(column) or "") for column in headers)
            for row in registry["records"]
        )
        user_prompt = f"""===== KEYWORD REFERENCE GUIDE =====
{guide}

===== TEMPORARY LEGACY KEYWORD REGISTRY =====
This is the only available registry and is explicitly a temporary legacy authority. Use its IDs exactly. Do not imply it is current.
{' | '.join(headers)}
{registry_text}

===== DOCUMENT =====
{header}

{source_body}

Return only SUMMARY and KEYWORDS in the required format."""
        request = _provider_request(
            request_id=request_id,
            stage_id="summary_keywords",
            stage_label="Summary & Keywords",
            document=document,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            max_output_tokens=max_output_tokens,
            accepted_upstream_stage_ids=[],
        )
        response = self.client.generate(request)
        validation = (
            validate_summary_keywords_output(response.raw_output, source_body=source_body, registry=registry)
            if response.status == ExecutionStatus.COMPLETED.value
            else {"valid": False, "issues": response.errors}
        )
        status = response.status if response.status != ExecutionStatus.COMPLETED.value else (
            ExecutionStatus.COMPLETED.value if validation["valid"] else ExecutionStatus.VALIDATION_FAILED.value
        )
        provenance = {
            "contract_version": SUMMARY_CONTRACT,
            "provider": response.provider,
            "model": response.model,
            "real_chatbot_execution": response.real_chatbot_execution,
            "request_id": request_id,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "keyword_registry": {key: registry[key] for key in ("authority_status", "filename", "sha256")},
            "prompt_character_count": len(system_prompt) + len(user_prompt),
            "model_call_completed": response.status == ExecutionStatus.COMPLETED.value,
        }
        return GenerationResult(status, response.raw_output, validation, provenance, validation,
                                response.provider, response.model, response.error, request)


class AssemblerGenerationService:
    def __init__(self, *, client: StageGenerationClient | None = None) -> None:
        self.client = client or stage_generation_client()

    def run(
        self,
        *,
        clause_output: StageOutput,
        entity_output: StageOutput,
        occurrence_output: StageOutput,
        request_id: str,
        conservation_test: bool = False,
        max_output_tokens: int = 2048,
    ) -> GenerationResult:
        if clause_output.status != StageOutput.Status.ACCEPTED:
            raise SummaryAssemblerError("Assembler requires accepted Clause output")
        if not entity_downstream_is_eligible(entity_output):
            raise SummaryAssemblerError("Assembler requires human-reviewed Entity output")
        if occurrence_output.status != StageOutput.Status.ACCEPTED and not conservation_test:
            raise SummaryAssemblerError("Assembler requires accepted Occurrence output")
        if len({clause_output.document_id, entity_output.document_id, occurrence_output.document_id}) != 1:
            raise SummaryAssemblerError("Assembler inputs belong to different documents")
        document = clause_output.document
        header, _ = _document_header_and_body(document)
        entity_package = build_entity_downstream_package(entity_output)
        system_prompt = ASSEMBLER_PROMPT.read_text(encoding="utf-8-sig")
        instructions = ASSEMBLER_INSTRUCTIONS.read_text(encoding="utf-8-sig")
        user_prompt = f"""===== GOVERNING ASSEMBLY INSTRUCTIONS =====
{instructions}

===== DOCUMENT HEADER =====
{header}

===== HUMAN-APPROVED ENTITY REGISTRY =====
{_entity_lines(entity_package)}

===== OCCURRENCES REGISTRY =====
{occurrence_output.raw_output}

Assemble only the clauses present in the Occurrences Registry. Reproduce every E:, A:, and Q: line character-for-character and in the same order. You may add only the entity, relationship, and temporal tags permitted by the governing instructions. Return no commentary."""
        request = _provider_request(
            request_id=request_id,
            stage_id="tag_assembler",
            stage_label="Tag Assembler",
            document=document,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            max_output_tokens=max_output_tokens,
            accepted_upstream_stage_ids=["clause_parser", "entity_registry"] + (
                ["occurrences_registry"] if occurrence_output.status == StageOutput.Status.ACCEPTED else []
            ),
        )
        response = self.client.generate(request)
        validation = (
            validate_occurrence_conservation(occurrence_output.raw_output, response.raw_output)
            if response.status == ExecutionStatus.COMPLETED.value
            else {"valid": False, "issues": response.errors}
        )
        status = response.status if response.status != ExecutionStatus.COMPLETED.value else (
            ExecutionStatus.COMPLETED.value if validation["valid"] else ExecutionStatus.VALIDATION_FAILED.value
        )
        provenance = {
            "contract_version": ASSEMBLER_CONTRACT,
            "provider": response.provider,
            "model": response.model,
            "real_chatbot_execution": response.real_chatbot_execution,
            "request_id": request_id,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_clause_stage_output_id": clause_output.pk,
            "source_entity_stage_output_id": entity_output.pk,
            "source_occurrence_stage_output_id": occurrence_output.pk,
            "source_occurrence_status": occurrence_output.status,
            "conservation_test": conservation_test,
            "production_eligible_input": occurrence_output.status == StageOutput.Status.ACCEPTED,
            "prompt_character_count": len(system_prompt) + len(user_prompt),
            "model_call_completed": response.status == ExecutionStatus.COMPLETED.value,
        }
        return GenerationResult(status, response.raw_output, {"validation": validation}, provenance,
                                validation, response.provider, response.model, response.error, request)


def _entity_lines(package: dict[str, Any]) -> str:
    return "\n".join(
        str(row.get("raw_line") or row.get("reviewed_raw_line") or row)
        for row in package.get("reviewed_rows", [])
    )


def _provider_request(
    *, request_id: str, stage_id: str, stage_label: str, document: Any,
    system_prompt: str, user_prompt: str, max_output_tokens: int,
    accepted_upstream_stage_ids: list[str],
) -> dict[str, Any]:
    return {
        "schema_version": "emtl-stage-execution-request-v1",
        "contract_version": "emtl-stage-contract-v1",
        "payload_schema_version": "emtl-provider-api-payload-draft-v1",
        "request_id": request_id,
        "stage_id": stage_id,
        "stage_label": stage_label,
        "provider": "gpu_local",
        "document_id": document.doc_id,
        "document_title": document.title,
        "document_type": document.document_type,
        "required_stage_ids": [],
        "accepted_upstream_stage_ids": accepted_upstream_stage_ids,
        "correction_requested": False,
        "inputs": {},
        "prompt_package": {
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
            "prompt_character_count": len(system_prompt) + len(user_prompt),
            "loaded_files": [],
            "source_completeness": {"source_complete": True, "truncation_applied": False},
        },
        "options": {"timeout_seconds": 3600, "max_output_tokens": max_output_tokens},
    }
