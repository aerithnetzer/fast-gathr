from __future__ import annotations

import csv
import hashlib
import io
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile


MANIFEST_FILENAME = "chatbot_bundle_manifest.json"
DEFAULT_KNOWLEDGE_CHAR_LIMIT = 120_000


def project_root() -> Path:
    return Path(__file__).resolve().parents[3]


def manifest_path() -> Path:
    return Path(__file__).resolve().parents[1] / MANIFEST_FILENAME


@dataclass(frozen=True)
class StageBundle:
    stage_id: str
    stage_label: str
    system_prompt_path: Path
    instruction_paths: tuple[Path, ...]
    example_paths: tuple[Path, ...]
    controlled_list_paths: tuple[Path, ...]
    boilerplate_paths: tuple[Path, ...]
    retrieval_resource_paths: tuple[Path, ...]
    input_requirements: tuple[str, ...]
    expected_output_type: str
    missing_files: tuple[str, ...]
    substitutions: tuple[dict[str, str], ...]
    system_prompt_resolution: dict[str, Any]
    old_files_not_used: tuple[str, ...]
    ambiguity_flags: tuple[str, ...]

    @property
    def knowledge_paths(self) -> tuple[Path, ...]:
        return (
            self.instruction_paths
            + self.example_paths
            + self.controlled_list_paths
            + self.boilerplate_paths
        )


class ChatbotBundleManifest:
    def __init__(self, path: str | Path | None = None) -> None:
        self.path = Path(path) if path else manifest_path()
        self.data = json.loads(self.path.read_text(encoding="utf-8"))

    def stage_ids(self) -> tuple[str, ...]:
        return tuple(self.data.get("stages", {}).keys())

    def get_stage(self, stage_id: str) -> StageBundle:
        stage = self.data.get("stages", {}).get(stage_id)
        if not isinstance(stage, dict):
            raise KeyError(f"Unknown chatbot stage: {stage_id}")
        root = project_root()

        def paths(key: str) -> tuple[Path, ...]:
            return tuple(root / str(value) for value in stage.get(key, []))

        return StageBundle(
            stage_id=stage_id,
            stage_label=str(stage.get("stage_label") or stage_id),
            system_prompt_path=root / str(stage.get("system_prompt") or ""),
            instruction_paths=paths("instructions"),
            example_paths=paths("examples"),
            controlled_list_paths=paths("controlled_lists"),
            boilerplate_paths=paths("boilerplate"),
            retrieval_resource_paths=paths("retrieval_resources"),
            input_requirements=tuple(str(value) for value in stage.get("input_requirements", [])),
            expected_output_type=str(stage.get("expected_output_type") or "text"),
            missing_files=tuple(str(value) for value in stage.get("missing_files", [])),
            substitutions=tuple(stage.get("substitutions", [])),
            system_prompt_resolution=dict(stage.get("system_prompt_resolution") or {}),
            old_files_not_used=tuple(str(value) for value in stage.get("old_files_not_used", [])),
            ambiguity_flags=tuple(str(value) for value in stage.get("ambiguity_flags", [])),
        )

    def audit(self) -> dict[str, Any]:
        stages: dict[str, Any] = {}
        for stage_id in self.stage_ids():
            bundle = self.get_stage(stage_id)
            selected_paths = (bundle.system_prompt_path,) + bundle.knowledge_paths
            stages[stage_id] = {
                "stage_label": bundle.stage_label,
                "selected_files": [str(path.relative_to(project_root())) for path in selected_paths],
                "selected_files_exist": {str(path.relative_to(project_root())): path.exists() for path in selected_paths},
                "declared_missing_files": list(bundle.missing_files),
                "substitutions": list(bundle.substitutions),
                "system_prompt_resolution": dict(bundle.system_prompt_resolution),
                "ambiguity_flags": list(bundle.ambiguity_flags),
            }
        return {
            "schema_version": self.data.get("schema_version"),
            "manifest_path": str(self.path),
            "stages": stages,
        }


class KnowledgeFileLoader:
    def __init__(self, char_limit: int = DEFAULT_KNOWLEDGE_CHAR_LIMIT) -> None:
        self.char_limit = max(1, int(char_limit))

    def read(self, path: Path) -> tuple[str, dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md", ".csv", ".json", ".rtf"}:
            text = path.read_text(encoding="utf-8-sig", errors="replace")
        elif suffix == ".docx":
            text = self._read_docx(path)
        elif suffix == ".xlsx":
            text = self._read_xlsx(path)
        else:
            raise ValueError(f"Unsupported chatbot knowledge file: {path.name}")
        truncated = len(text) > self.char_limit
        return text[: self.char_limit], {
            "path": str(path.relative_to(project_root())),
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            "characters_loaded": min(len(text), self.char_limit),
            "characters_available": len(text),
            "truncated": truncated,
        }

    def _read_docx(self, path: Path) -> str:
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        with ZipFile(path) as archive:
            root = ET.fromstring(archive.read("word/document.xml"))
        paragraphs: list[str] = []
        for paragraph in root.findall(".//w:p", namespace):
            text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace))
            if text.strip():
                paragraphs.append(text)
        return "\n".join(paragraphs)

    def _read_xlsx(self, path: Path) -> str:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        stream = io.StringIO()
        writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
        for worksheet in workbook.worksheets:
            writer.writerow([f"SHEET: {worksheet.title}"])
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else str(value) for value in row])
        return stream.getvalue()
