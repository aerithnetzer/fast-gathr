from __future__ import annotations

import re
from difflib import SequenceMatcher
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

EVENT_LOOKUP_BACKEND = "local_tfidf"
EVENT_LOOKUP_SCORE_TYPE = "tfidf_cosine_with_lexical_boost"
DETERMINISTIC_LEXICAL_BACKEND = "deterministic_lexical_authority"
DETERMINISTIC_LEXICAL_SCORE_TYPE = "sequence_token_overlap_v1"
EVENT_SHEET_NAME = "19. Events"
MASTER_LIST_FILENAME = "Master List_Dec_3_copy.xlsx"
VECTOR_EVENT_LIST_RELATIVE_PATH = Path("Chatbot docs") / "Events_List_VectorLLM_v1.xlsx"


@dataclass(frozen=True)
class EventVocabularyEntry:
    event_id: str
    headword: str
    classification: str
    definition: str
    keywords: str
    vector_examples: str
    llm_examples: str
    source: str
    source_row: int | None = None


@dataclass(frozen=True)
class ScoredEventCandidate:
    entry: EventVocabularyEntry
    score: float
    match_reason: str


class EventVocabularyProvider(Protocol):
    def load_entries(self) -> list[EventVocabularyEntry]:
        """Return canonical Event vocabulary entries available to this backend."""


class SimilarityBackend(Protocol):
    backend_name: str
    score_type: str

    def rank(
        self,
        query_text: str,
        entries: list[EventVocabularyEntry],
        text_builder: "EventCandidateTextBuilder",
        top_k: int,
    ) -> list[ScoredEventCandidate]:
        """Rank Event entries against an open trigger/clause query."""


def default_master_list_path() -> Path:
    return Path(__file__).resolve().parents[3] / MASTER_LIST_FILENAME


def default_vector_event_list_path() -> Path:
    return Path(__file__).resolve().parents[3] / VECTOR_EVENT_LIST_RELATIVE_PATH


class ExcelEventVocabularyProvider:
    """Current local provider backed by the root Master List workbook."""

    def __init__(
        self,
        workbook_path: str | Path | None = None,
        sheet_name: str = EVENT_SHEET_NAME,
    ) -> None:
        self.workbook_path = Path(workbook_path) if workbook_path else default_master_list_path()
        self.sheet_name = sheet_name

    def load_entries(self) -> list[EventVocabularyEntry]:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        worksheet = workbook[self.sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        indexes = {header: index for index, header in enumerate(headers) if header}
        required = {"ID", "Element"}
        missing = sorted(required - set(indexes))
        if missing:
            raise ValueError(f"Event sheet missing required column(s): {', '.join(missing)}")

        source = f"{self.workbook_path.name}:{self.sheet_name}"
        entries: list[EventVocabularyEntry] = []
        for row in rows[1:]:
            event_id = _cell(row, indexes, "ID")
            headword = _cell(row, indexes, "Element")
            if not event_id.startswith("E-") or not headword:
                continue
            entries.append(
                EventVocabularyEntry(
                    event_id=event_id,
                    headword=headword,
                    classification=_cell(row, indexes, "Classification"),
                    definition=_cell(row, indexes, "Definition"),
                    keywords=_cell(row, indexes, "Keywords"),
                    vector_examples="",
                    llm_examples="",
                    source=source,
                    source_row=None,
                )
            )
        return entries


class VectorWorkbookEventVocabularyProvider:
    """Current Event authority with separate retrieval and LLM-facing examples."""

    def __init__(self, workbook_path: str | Path | None = None) -> None:
        self.workbook_path = Path(workbook_path) if workbook_path else default_vector_event_list_path()

    def load_entries(self) -> list[EventVocabularyEntry]:
        from openpyxl import load_workbook

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        indexes = {header: index for index, header in enumerate(headers) if header}
        required = {"ID", "Headword", "Definition", "Vector Example"}
        missing = sorted(required - set(indexes))
        if missing:
            raise ValueError(f"Vector Event list missing required column(s): {', '.join(missing)}")

        source = f"{self.workbook_path.name}:{worksheet.title}"
        entries: list[EventVocabularyEntry] = []
        for excel_row, row in enumerate(rows[1:], start=2):
            event_id = _cell(row, indexes, "ID")
            headword = _cell(row, indexes, "Headword")
            if not event_id.startswith("E-") or not headword:
                continue
            entries.append(
                EventVocabularyEntry(
                    event_id=event_id,
                    headword=headword,
                    classification="",
                    definition=_cell(row, indexes, "Definition"),
                    keywords="",
                    vector_examples=_cell(row, indexes, "Vector Example"),
                    llm_examples=_cell(row, indexes, "LLM Example"),
                    source=source,
                    source_row=excel_row,
                )
            )
        return entries


class EventCandidateTextBuilder:
    """Build current retrieval text: Headword + Definition + Vector Example."""

    def build(self, entry: EventVocabularyEntry) -> str:
        parts = [
            f"Headword: {entry.headword}",
        ]
        if entry.definition:
            parts.append(f"Definition: {entry.definition}")
        if entry.vector_examples:
            parts.append(f"Vector Example: {entry.vector_examples}")
        return ". ".join(parts)

    def lexical_text(self, entry: EventVocabularyEntry) -> str:
        return " ".join(
            value
            for value in (
                entry.headword,
                entry.definition,
                entry.vector_examples,
            )
            if value
        )


class TfidfSimilarityBackend:
    """Banned legacy backend retained only to make accidental use fail closed."""

    backend_name = EVENT_LOOKUP_BACKEND
    score_type = EVENT_LOOKUP_SCORE_TYPE

    def rank(
        self,
        query_text: str,
        entries: list[EventVocabularyEntry],
        text_builder: EventCandidateTextBuilder,
        top_k: int,
    ) -> list[ScoredEventCandidate]:
        raise RuntimeError("TF-IDF Event lookup is banned; no fallback is permitted.")


class DeterministicLexicalSimilarityBackend:
    """Model-free contract backend; explicitly not dense/vector retrieval."""

    backend_name = DETERMINISTIC_LEXICAL_BACKEND
    score_type = DETERMINISTIC_LEXICAL_SCORE_TYPE

    def rank(self, query_text, entries, text_builder, top_k):
        query = str(query_text or "").strip()
        if not query or top_k <= 0:
            return []
        query_norm = _normalize_phrase(query)
        query_tokens = set(_normalized_tokens(query))
        scored = []
        for entry in entries:
            authority_text = text_builder.lexical_text(entry)
            authority_norm = _normalize_phrase(authority_text)
            authority_tokens = set(_normalized_tokens(authority_text))
            sequence = SequenceMatcher(None, query_norm, authority_norm).ratio()
            overlap = len(query_tokens & authority_tokens) / max(len(query_tokens), 1)
            exact_boost, reason = _lexical_boost(query, entry, text_builder)
            score = min(1.0, (sequence * 0.45) + (overlap * 0.55) + exact_boost)
            scored.append(ScoredEventCandidate(
                entry=entry,
                score=score,
                match_reason=(
                    f"sequence={sequence:.3f}; token_overlap={overlap:.3f}; "
                    f"authority_match={reason or 'none'}"
                ),
            ))
        scored.sort(key=lambda candidate: (-candidate.score, candidate.entry.event_id))
        return scored[:top_k]


class EventLookupService:
    def __init__(
        self,
        vocabulary_provider: EventVocabularyProvider,
        text_builder: EventCandidateTextBuilder | None = None,
        similarity_backend: SimilarityBackend | None = None,
    ) -> None:
        self.vocabulary_provider = vocabulary_provider
        self.text_builder = text_builder or EventCandidateTextBuilder()
        self.similarity_backend = similarity_backend or TfidfSimilarityBackend()
        self._entries: list[EventVocabularyEntry] | None = None

    def lookup(self, query_text: str, top_k: int = 5) -> list[dict[str, object]]:
        entries = self._load_entries()
        candidates = self.similarity_backend.rank(
            query_text=query_text,
            entries=entries,
            text_builder=self.text_builder,
            top_k=top_k,
        )
        return [self._candidate_to_dict(candidate) for candidate in candidates]

    def _load_entries(self) -> list[EventVocabularyEntry]:
        if self._entries is None:
            self._entries = self.vocabulary_provider.load_entries()
        return self._entries

    def _candidate_to_dict(self, candidate: ScoredEventCandidate) -> dict[str, object]:
        entry = candidate.entry
        return {
            "event_id": entry.event_id,
            "headword": entry.headword,
            "classification": entry.classification,
            "definition": entry.definition,
            "keywords": entry.keywords,
            "vector_examples": _split_examples(entry.vector_examples),
            "llm_examples": _split_examples(entry.llm_examples),
            "matched_example": _best_matching_example(entry.vector_examples, candidate),
            "score": round(candidate.score, 6),
            "match_reason": candidate.match_reason,
            "source": entry.source,
            "source_row": entry.source_row,
            "backend": self.similarity_backend.backend_name,
            "score_type": self.similarity_backend.score_type,
            "development_only": True,
        }


def build_default_event_lookup_service() -> EventLookupService:
    return EventLookupService(
        vocabulary_provider=VectorWorkbookEventVocabularyProvider(),
        similarity_backend=DeterministicLexicalSimilarityBackend(),
    )


def _split_examples(value: str) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def _best_matching_example(value: str, candidate: ScoredEventCandidate) -> str:
    examples = _split_examples(value)
    return examples[0] if examples else ""


def _cell(row: tuple[object, ...], indexes: dict[str, int], column: str) -> str:
    index = indexes.get(column)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


STOP_TOKENS = {
    "a",
    "an",
    "and",
    "be",
    "by",
    "did",
    "for",
    "from",
    "he",
    "her",
    "his",
    "in",
    "of",
    "or",
    "said",
    "same",
    "she",
    "that",
    "the",
    "to",
    "was",
    "were",
    "with",
    "away",
}


def _lexical_boost(
    query: str,
    entry: EventVocabularyEntry,
    text_builder: EventCandidateTextBuilder,
) -> tuple[float, str]:
    query_norm = _normalize_phrase(query)
    canonical_norm = _normalize_phrase(" ".join([entry.headword, entry.vector_examples]))
    scope_norm = _normalize_phrase(entry.definition)
    if not query_norm or not (canonical_norm or scope_norm):
        return 0.0, ""

    if query_norm == _normalize_phrase(entry.headword):
        return 0.30, "exact headword match"
    if query_norm and query_norm in _normalize_phrase(entry.vector_examples):
        return 0.28, "query phrase appears in vector examples"
    if _normalize_phrase(entry.headword) in query_norm:
        return 0.20, "headword appears in query"

    query_tokens = set(_normalized_tokens(query))
    canonical_tokens = set(_normalized_tokens(" ".join([entry.headword, entry.vector_examples])))
    canonical_overlap = sorted(query_tokens & canonical_tokens)
    if canonical_overlap:
        overlap_ratio = len(canonical_overlap) / max(len(query_tokens), 1)
        boost = min(0.35, 0.12 + (overlap_ratio * 0.34))
        return boost, "headword/vector-example token overlap: " + ", ".join(canonical_overlap[:6])

    scope_tokens = set(_normalized_tokens(entry.definition))
    overlap = sorted(query_tokens & scope_tokens)
    if not overlap:
        return 0.0, ""
    overlap_ratio = len(overlap) / max(len(query_tokens), 1)
    boost = min(0.18, 0.04 + (overlap_ratio * 0.18))
    return boost, "definition token overlap: " + ", ".join(overlap[:6])


def _normalized_tokens(text: str) -> list[str]:
    tokens = re.findall(r"[a-zA-Z0-9]+", _normalize_uv(text))
    normalized: list[str] = []
    for token in tokens:
        root = _token_root(token)
        if root and root not in STOP_TOKENS and len(root) > 1:
            normalized.append(root)
    return normalized


def _normalize_phrase(text: str) -> str:
    return " ".join(_normalized_tokens(text))


def _normalize_uv(text: str) -> str:
    lowered = str(text or "").lower()
    return lowered.replace("v", "u")


def _token_root(token: str) -> str:
    token = token.lower().strip()
    if not token:
        return ""
    if token.startswith(("pay", "pai")):
        return "pay"
    if token.startswith(("deliuer", "deliuer", "delyuer", "delyver", "deliver")):
        return "deliver"
    if token.startswith(("forbear", "forbeare", "forbere", "forborn")):
        return "forbear"
    if token.startswith(("bound", "bonde", "bond", "band", "bande")):
        return "bond"
    if token.startswith(("utter", "vtter")):
        return "utter"
    if token.startswith(("sell", "sale", "sold", "sould")):
        return "sale"
    if token.startswith("put"):
        return "put"
    if token.endswith("inge"):
        token = token[:-4]
    elif token.endswith("ing"):
        token = token[:-3]
    elif token.endswith("ed"):
        token = token[:-2]
    elif token.endswith("de"):
        token = token[:-2]
    elif token.endswith("es"):
        token = token[:-2]
    elif token.endswith("s") and len(token) > 3:
        token = token[:-1]
    return token
