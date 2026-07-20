from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .event_lookup import (
    DeterministicLexicalSimilarityBackend,
    EventCandidateTextBuilder,
    EventVocabularyEntry,
    ScoredEventCandidate,
    TfidfSimilarityBackend,
    default_vector_event_list_path,
)


CONTRACT_VERSION = "event-lookup-candidate-package-v1"
INDEX_VERSION = "event-authority-row-index-v1"
REQUIRED_COLUMNS = ("ID", "Headword", "Definition", "Vector Example", "LLM Example")


@dataclass(frozen=True)
class EventCutInput:
    event_cut_id: str
    document_id: str
    event_cut_text: str
    clause_id: str = ""
    clause_label: str = ""
    clause_text: str = ""
    trigger: str = ""
    source: str = "command"
    source_offsets: dict[str, int] | None = None
    provenance: dict[str, Any] = field(default_factory=dict)

    def validate(self) -> None:
        if not self.event_cut_id.strip():
            raise ValueError("event_cut_id is required.")
        if not self.event_cut_text.strip():
            raise ValueError("event_cut_text is required.")
        if self.source not in {"manual", "llm", "fixture", "command"}:
            raise ValueError("source must be manual, llm, fixture, or command.")

    def as_dict(self) -> dict[str, Any]:
        return {
            "event_cut_id": self.event_cut_id,
            "document_id": self.document_id,
            "clause_id": self.clause_id,
            "clause_label": self.clause_label,
            "clause_text": self.clause_text,
            "trigger": self.trigger,
            "event_cut_text": self.event_cut_text,
            "source": self.source,
            "source_offsets": self.source_offsets,
            "provenance": dict(self.provenance),
        }


@dataclass(frozen=True)
class EventRegistrySnapshot:
    entries: tuple[EventVocabularyEntry, ...]
    issues: tuple[dict[str, Any], ...]
    resource_path: str
    workbook_sheet: str
    resource_hash: str
    registry_version: str
    total_data_rows: int
    indexed_row_count: int


def load_event_registry_snapshot(workbook_path: Path | None = None) -> EventRegistrySnapshot:
    selected = (workbook_path or default_vector_event_list_path()).resolve()
    authority = default_vector_event_list_path().resolve()
    if workbook_path is None and selected != authority:
        raise ValueError("Authoritative Event workbook resolution is ambiguous.")
    digest = hashlib.sha256(selected.read_bytes()).hexdigest()
    workbook = load_workbook(selected, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet_title = worksheet.title
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError("Authoritative Event workbook is empty.")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Event workbook missing required columns: {', '.join(missing)}")
    indexes = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    issues: list[dict[str, Any]] = []
    entries: list[EventVocabularyEntry] = []
    id_rows: dict[str, list[int]] = {}
    headword_rows: dict[str, list[int]] = {}
    data_rows = 0
    for excel_row, row in enumerate(rows[1:], start=2):
        values = ["" if value is None else str(value).strip() for value in row]
        if not any(values):
            continue
        data_rows += 1
        value = lambda name: values[indexes[name]] if indexes[name] < len(values) else ""
        event_id, headword = value("ID"), value("Headword")
        if not event_id or not headword:
            issues.append({
                "code": "empty_authority_id_or_headword",
                "severity": "error",
                "source_row": excel_row,
                "event_id_present": bool(event_id),
                "headword_present": bool(headword),
                "indexed": False,
            })
            continue
        id_rows.setdefault(event_id, []).append(excel_row)
        headword_rows.setdefault(headword.casefold(), []).append(excel_row)
        entries.append(EventVocabularyEntry(
            event_id=event_id,
            headword=headword,
            classification="",
            definition=value("Definition"),
            keywords="",
            vector_examples=value("Vector Example"),
            llm_examples=value("LLM Example"),
            source=f"{selected.name}:{worksheet_title}",
            source_row=excel_row,
        ))
    for code, groups in (("duplicate_event_id", id_rows), ("duplicate_headword", headword_rows)):
        for authority_value, source_rows in groups.items():
            if len(source_rows) > 1:
                issues.append({
                    "code": code,
                    "severity": "ambiguity",
                    "authority_value": authority_value,
                    "source_rows": source_rows,
                })
    return EventRegistrySnapshot(
        entries=tuple(entries),
        issues=tuple(issues),
        resource_path=selected.as_posix(),
        workbook_sheet=worksheet_title,
        resource_hash=digest,
        registry_version=f"sha256:{digest}",
        total_data_rows=data_rows,
        indexed_row_count=len(entries),
    )


class EventLookupPackageBuilder:
    def __init__(self, *, backend=None, workbook_path: Path | None = None) -> None:
        self.backend = backend or DeterministicLexicalSimilarityBackend()
        self.workbook_path = workbook_path
        backend_name = str(getattr(self.backend, "backend_name", "")).lower()
        if isinstance(self.backend, TfidfSimilarityBackend) or "tfidf" in backend_name:
            raise ValueError("TF-IDF Event lookup is banned; no fallback is permitted.")

    def build(self, event_cut: EventCutInput, *, top_k: int = 20) -> dict[str, Any]:
        event_cut.validate()
        if top_k <= 0:
            raise ValueError("top_k must be positive.")
        snapshot = load_event_registry_snapshot(self.workbook_path)
        query_text = event_cut.event_cut_text.strip()
        text_builder = EventCandidateTextBuilder()
        ranked: list[ScoredEventCandidate] = self.backend.rank(
            query_text, list(snapshot.entries), text_builder, top_k
        )
        package_seed = json.dumps({
            "event_cut": event_cut.as_dict(),
            "registry_version": snapshot.registry_version,
            "backend": self.backend.backend_name,
            "top_k": top_k,
        }, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        package_id = "event-lookup-" + hashlib.sha256(package_seed.encode("utf-8")).hexdigest()[:24]
        candidates = [self._candidate(candidate, rank, snapshot) for rank, candidate in enumerate(ranked, 1)]
        return {
            "contract_version": CONTRACT_VERSION,
            "package_id": package_id,
            "event_cut": event_cut.as_dict(),
            "query_text": query_text,
            "top_k_requested": top_k,
            "registry_version": snapshot.registry_version,
            "registry_hashes": {Path(snapshot.resource_path).name: snapshot.resource_hash},
            "resource_path": snapshot.resource_path,
            "workbook_sheet": snapshot.workbook_sheet,
            "backend_name": self.backend.backend_name,
            "backend_version": INDEX_VERSION,
            "backend_is_dense_vector": False,
            "tfidf_used": False,
            "registry_validation": {
                "total_data_rows": snapshot.total_data_rows,
                "indexed_row_count": snapshot.indexed_row_count,
                "issues": list(snapshot.issues),
            },
            "candidates": candidates,
        }

    def _candidate(self, candidate, rank, snapshot):
        entry = candidate.entry
        return {
            "rank": rank,
            "event_id": entry.event_id,
            "headword": entry.headword,
            "classification": entry.classification,
            "definition": entry.definition,
            "vector_example": entry.vector_examples,
            "llm_example": entry.llm_examples,
            "score": round(candidate.score, 6),
            "score_type": self.backend.score_type,
            "match_reason": candidate.match_reason,
            "source_workbook": Path(snapshot.resource_path).name,
            "source_sheet": snapshot.workbook_sheet,
            "source_row": entry.source_row,
            "provenance": {
                "resource_hash": snapshot.resource_hash,
                "registry_version": snapshot.registry_version,
                "authority_pair_immutable": True,
            },
            "development_only": True,
        }


def write_event_lookup_package(package: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")
