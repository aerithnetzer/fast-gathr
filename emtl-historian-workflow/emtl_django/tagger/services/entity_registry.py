from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Protocol

from .chatbot_bundle import project_root


REGISTRY_CONTRACT_VERSION = "emtl-entity-registry-v1"
MAX_CANDIDATE_LIMIT = 50


@dataclass(frozen=True)
class RegistryResourceSpec:
    name: str
    relative_path: Path
    sha256: str
    kind: str


ENTITY_REGISTRY_RESOURCES = (
    RegistryResourceSpec(
        name="Entity_List.xlsx",
        relative_path=Path("Chatbot docs") / "Claude chatbots" / "Entity_List.xlsx",
        sha256="4e19066d63f24da56545898f75c271f148f8661e0dd48b6e55a8be8a7ff4a28d",
        kind="entity",
    ),
    RegistryResourceSpec(
        name="Disambiguation_Watch_List.csv",
        relative_path=Path("Chatbot docs") / "Claude chatbots" / "Disambiguation_Watch_List.csv",
        sha256="e4f6f2f906ca6977d44df3c0a997de167b7bc18908ed6ab35cbfd0fc9ac3c05e",
        kind="watch",
    ),
    RegistryResourceSpec(
        name="Attributes_List.csv",
        relative_path=Path("Chatbot docs") / "Claude chatbots" / "Attributes_List.csv",
        sha256="8c88368e24ca8bc1ee7856b681214d9b49e86acf6e7882c0288de214e9c8fb4f",
        kind="attribute",
    ),
)


class RegistryIntegrityError(RuntimeError):
    pass


def normalize_registry_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", " ", normalized).strip().casefold()


@dataclass(frozen=True)
class RegistryField:
    column: int
    name: str
    value: str

    def as_dict(self) -> dict[str, Any]:
        return {"column": self.column, "name": self.name, "value": self.value}


@dataclass(frozen=True)
class RegistryProvenance:
    source_file: str
    source_path: str
    source_hash: str
    registry_version: str
    sheet: str = ""
    row: int | None = None
    row_end: int | None = None

    def as_dict(self) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "source_path": self.source_path,
            "source_hash": self.source_hash,
            "registry_version": self.registry_version,
            "sheet": self.sheet,
            "row": self.row,
            "row_end": self.row_end,
        }


@dataclass(frozen=True)
class RegistryRecord:
    kind: str
    record_id: str
    headword: str
    normalized_headword: str
    fields: tuple[RegistryField, ...]
    provenance: RegistryProvenance

    def values(self, field_name: str) -> tuple[str, ...]:
        wanted = normalize_registry_name(field_name)
        return tuple(
            field.value
            for field in self.fields
            if normalize_registry_name(field.name) == wanted and field.value
        )

    def as_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "id": self.record_id,
            "headword": self.headword,
            "normalized_headword": self.normalized_headword,
            "fields": [field.as_dict() for field in self.fields],
            "provenance": self.provenance.as_dict(),
        }


@dataclass(frozen=True)
class RegistryIssue:
    code: str
    message: str
    severity: str = "ambiguity"
    record_id: str = ""
    headword: str = ""
    provenance: tuple[RegistryProvenance, ...] = ()
    details: dict[str, Any] = field(default_factory=dict)

    def as_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
            "id": self.record_id,
            "headword": self.headword,
            "provenance": [item.as_dict() for item in self.provenance],
            "details": self.details,
        }


@dataclass(frozen=True)
class LookupResult:
    query: str
    normalized_query: str
    status: str
    match_type: str
    candidates: tuple[RegistryRecord, ...]
    ambiguities: tuple[RegistryIssue, ...]
    registry_version: str
    resource_hashes: dict[str, str]

    def as_dict(self) -> dict[str, Any]:
        return {
            "query": self.query,
            "normalized_query": self.normalized_query,
            "status": self.status,
            "match_type": self.match_type,
            "candidates": [item.as_dict() for item in self.candidates],
            "ambiguities": [item.as_dict() for item in self.ambiguities],
            "registry_version": self.registry_version,
            "resource_hashes": dict(self.resource_hashes),
        }


@dataclass(frozen=True)
class VectorCandidate:
    record_id: str
    score: float
    metadata: dict[str, Any] = field(default_factory=dict)


class OptionalVectorCandidateLayer(Protocol):
    layer_name: str

    def query(self, text: str, top_k: int) -> Iterable[VectorCandidate]:
        """Return candidate IDs only. Registry validation remains authoritative."""


@dataclass(frozen=True)
class CandidateMatch:
    record: RegistryRecord
    score: float
    layer: str
    matched_text: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "record": self.record.as_dict(),
            "score": self.score,
            "layer": self.layer,
            "matched_text": self.matched_text,
        }


@dataclass(frozen=True)
class CandidateRetrievalResult:
    query: str
    normalized_query: str
    top_k: int
    candidates: tuple[CandidateMatch, ...]
    ambiguities: tuple[RegistryIssue, ...]
    layers_used: tuple[str, ...]
    registry_version: str
    resource_hashes: dict[str, str]
    final_id_selected: bool = False

    def as_dict(self) -> dict[str, Any]:
        return {
            "query": self.query,
            "normalized_query": self.normalized_query,
            "top_k": self.top_k,
            "candidates": [item.as_dict() for item in self.candidates],
            "ambiguities": [item.as_dict() for item in self.ambiguities],
            "layers_used": list(self.layers_used),
            "registry_version": self.registry_version,
            "resource_hashes": dict(self.resource_hashes),
            "final_id_selected": self.final_id_selected,
        }


@dataclass(frozen=True)
class RegistrySnapshot:
    registry_version: str
    resource_hashes: dict[str, str]
    entity_records: tuple[RegistryRecord, ...]
    attribute_records: tuple[RegistryRecord, ...]
    watch_records: tuple[RegistryRecord, ...]
    issues: tuple[RegistryIssue, ...]

    def statistics(self) -> dict[str, Any]:
        issue_counts: dict[str, int] = defaultdict(int)
        for issue in self.issues:
            issue_counts[issue.code] += 1
        entity_ids = [record.record_id for record in self.entity_records if record.record_id]
        attribute_ids = [record.record_id for record in self.attribute_records if record.record_id]
        return {
            "registry_version": self.registry_version,
            "resource_hashes": dict(self.resource_hashes),
            "entity_records": len(self.entity_records),
            "entity_unique_ids": len(set(entity_ids)),
            "attribute_records": len(self.attribute_records),
            "attribute_unique_ids": len(set(attribute_ids)),
            "watch_records": len(self.watch_records),
            "issues_by_code": dict(sorted(issue_counts.items())),
        }


class EntityRegistryService:
    """Read-only, source-faithful registry built from the three platform resources."""

    def __init__(
        self,
        *,
        root: str | Path | None = None,
        resources: tuple[RegistryResourceSpec, ...] = ENTITY_REGISTRY_RESOURCES,
    ) -> None:
        self.root = Path(root) if root is not None else project_root()
        self.resources = resources
        self._snapshot: RegistrySnapshot | None = None
        self._id_index: dict[str, tuple[RegistryRecord, ...]] = {}
        self._attribute_id_index: dict[str, tuple[RegistryRecord, ...]] = {}
        self._headword_index: dict[str, tuple[RegistryRecord, ...]] = {}
        self._watch_index: dict[str, tuple[RegistryRecord, ...]] = {}
        self._entity_search_terms: dict[tuple[str, str, int | None], tuple[str, ...]] = {}

    @property
    def snapshot(self) -> RegistrySnapshot:
        if self._snapshot is None:
            self._load()
        assert self._snapshot is not None
        return self._snapshot

    @property
    def registry_version(self) -> str:
        return self.snapshot.registry_version

    @property
    def resource_hashes(self) -> dict[str, str]:
        return dict(self.snapshot.resource_hashes)

    def lookup_id(self, record_id: str) -> LookupResult:
        self.snapshot
        query = str(record_id or "").strip()
        candidates = self._id_index.get(query, ())
        issues = self._issues_for_candidates(candidates, record_id=query)
        if not query:
            issues = (
                RegistryIssue(code="empty_id_query", message="ID query is empty.", severity="error"),
            )
            status = "invalid"
        elif not candidates:
            issues = (
                RegistryIssue(
                    code="id_not_found",
                    message=f"ID {query} is not present in Entity_List.xlsx.",
                    severity="error",
                    record_id=query,
                ),
            )
            status = "not_found"
        elif issues or len(candidates) != 1:
            status = "ambiguous"
        else:
            status = "match"
        return self._lookup_result(query, query, status, "id", candidates, issues)

    def lookup_headword(self, headword: str) -> LookupResult:
        self.snapshot
        query = str(headword or "")
        normalized = normalize_registry_name(query)
        candidates = self._headword_index.get(normalized, ())
        issues = self._issues_for_candidates(candidates, headword=query)
        if not normalized:
            issues = (
                RegistryIssue(code="empty_headword_query", message="Headword query is empty.", severity="error"),
            )
            status = "invalid"
        elif not candidates:
            status = "not_found"
        elif len(candidates) > 1 or issues:
            issues = issues + (
                RegistryIssue(
                    code="headword_has_multiple_candidates",
                    message=f"Normalized Headword {normalized!r} resolves to multiple records.",
                    headword=query,
                    provenance=tuple(record.provenance for record in candidates),
                    details={"candidate_count": len(candidates)},
                ),
            ) if len(candidates) > 1 else issues
            status = "ambiguous"
        else:
            status = "match"
        return self._lookup_result(query, normalized, status, "normalized_headword", candidates, issues)

    def validate_id_headword(self, record_id: str, headword: str) -> LookupResult:
        result = self.lookup_id(record_id)
        if result.status != "match":
            return result
        record = result.candidates[0]
        normalized = normalize_registry_name(headword)
        if normalized == record.normalized_headword and str(headword).strip() == record.headword:
            return self._lookup_result(
                str(headword), normalized, "match", "id_and_headword", result.candidates, ()
            )
        code = (
            "headword_string_difference"
            if normalized == record.normalized_headword
            else "id_headword_mismatch"
        )
        issue = RegistryIssue(
            code=code,
            message=(
                f"ID {record_id} resolves to Headword {record.headword!r}, not {headword!r}."
            ),
            severity="error",
            record_id=record_id,
            headword=headword,
            provenance=(record.provenance,),
            details={"registry_headword": record.headword},
        )
        return self._lookup_result(
            str(headword), normalized, "ambiguous", "id_and_headword", result.candidates, (issue,)
        )

    def validate_attribute_id(self, attribute_id: str, headword: str = "") -> LookupResult:
        self.snapshot
        query = str(attribute_id or "").strip()
        candidates = self._attribute_id_index.get(query, ())
        issues = self._issues_for_candidates(candidates, record_id=query)
        if not candidates:
            issues = (
                RegistryIssue(
                    code="attribute_id_not_found",
                    message=f"Attribute ID {query or '<empty>'} is not present in Attributes_List.csv.",
                    severity="error",
                    record_id=query,
                ),
            )
            status = "not_found" if query else "invalid"
        elif len(candidates) != 1 or issues:
            status = "ambiguous"
        elif headword and normalize_registry_name(headword) != candidates[0].normalized_headword:
            issues = (
                RegistryIssue(
                    code="attribute_id_headword_mismatch",
                    message=(
                        f"Attribute ID {query} resolves to {candidates[0].headword!r}, not {headword!r}."
                    ),
                    severity="error",
                    record_id=query,
                    headword=headword,
                    provenance=(candidates[0].provenance,),
                ),
            )
            status = "ambiguous"
        else:
            status = "match"
        return self._lookup_result(query, normalize_registry_name(headword or query), status, "attribute_id", candidates, issues)

    def lookup_watch(self, value: str) -> LookupResult:
        self.snapshot
        query = str(value or "")
        normalized = normalize_registry_name(query)
        candidates = self._watch_index.get(normalized, ())
        issues: tuple[RegistryIssue, ...] = ()
        if not normalized:
            status = "invalid"
        elif not candidates:
            status = "not_found"
        elif len(candidates) > 1:
            status = "ambiguous"
            issues = (
                RegistryIssue(
                    code="watch_disambiguation_required",
                    message=f"Watch term {query!r} has multiple listed senses.",
                    headword=query,
                    provenance=tuple(record.provenance for record in candidates),
                    details={"candidate_count": len(candidates)},
                ),
            )
        else:
            status = "match"
        return self._lookup_result(query, normalized, status, "watch_term", candidates, issues)

    def retrieve_candidates(
        self,
        query: str,
        *,
        top_k: int = 10,
        min_fuzzy_score: float = 0.55,
        vector_layer: OptionalVectorCandidateLayer | None = None,
    ) -> CandidateRetrievalResult:
        self.snapshot
        bounded_top_k = max(1, min(int(top_k), MAX_CANDIDATE_LIMIT))
        raw_query = str(query or "").strip()
        normalized_query = normalize_registry_name(raw_query)
        matches: dict[tuple[str, str, int | None], CandidateMatch] = {}
        ambiguities: list[RegistryIssue] = []
        layers = ["exact", "normalized", "lexical_fuzzy"]

        if normalized_query:
            for record in self.snapshot.entity_records:
                key = self._record_key(record)
                terms = self._entity_search_terms.get(key, ())
                best: CandidateMatch | None = None
                for term in terms:
                    normalized_term = normalize_registry_name(term)
                    if raw_query == term:
                        candidate = CandidateMatch(record, 1.0, "exact", term)
                    elif normalized_query == normalized_term:
                        candidate = CandidateMatch(record, 0.99, "normalized", term)
                    else:
                        score = SequenceMatcher(None, normalized_query, normalized_term).ratio()
                        if score < min_fuzzy_score:
                            continue
                        candidate = CandidateMatch(record, score, "lexical_fuzzy", term)
                    if best is None or self._candidate_sort_key(candidate) < self._candidate_sort_key(best):
                        best = candidate
                if best is not None:
                    matches[key] = best

        if vector_layer is not None and raw_query:
            layers.append(str(getattr(vector_layer, "layer_name", "optional_vector")))
            for vector_candidate in vector_layer.query(raw_query, bounded_top_k):
                validation = self.lookup_id(vector_candidate.record_id)
                if validation.status != "match":
                    ambiguities.extend(validation.ambiguities)
                    continue
                record = validation.candidates[0]
                key = self._record_key(record)
                candidate = CandidateMatch(
                    record=record,
                    score=max(0.0, min(float(vector_candidate.score), 1.0)),
                    layer=str(getattr(vector_layer, "layer_name", "optional_vector")),
                    matched_text=str(vector_candidate.metadata.get("matched_text") or ""),
                )
                existing = matches.get(key)
                if existing is None or self._candidate_sort_key(candidate) < self._candidate_sort_key(existing):
                    matches[key] = candidate

        ranked = sorted(matches.values(), key=self._candidate_sort_key)[:bounded_top_k]
        for candidate in ranked:
            ambiguities.extend(self._issues_for_candidates((candidate.record,), record_id=candidate.record.record_id))
        return CandidateRetrievalResult(
            query=raw_query,
            normalized_query=normalized_query,
            top_k=bounded_top_k,
            candidates=tuple(ranked),
            ambiguities=tuple(_deduplicate_issues(ambiguities)),
            layers_used=tuple(layers),
            registry_version=self.registry_version,
            resource_hashes=self.resource_hashes,
            final_id_selected=False,
        )

    def _load(self) -> None:
        actual_hashes: dict[str, str] = {}
        paths: dict[str, Path] = {}
        for resource in self.resources:
            path = self.root / resource.relative_path
            if not path.is_file():
                raise RegistryIntegrityError(f"Missing built-in registry resource: {path}")
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            if digest != resource.sha256:
                raise RegistryIntegrityError(
                    f"Hash mismatch for {resource.name}: expected {resource.sha256}, received {digest}"
                )
            actual_hashes[resource.name] = digest
            paths[resource.name] = path

        version_payload = {
            "contract": REGISTRY_CONTRACT_VERSION,
            "resources": {name: actual_hashes[name] for name in sorted(actual_hashes)},
        }
        version_digest = hashlib.sha256(
            json.dumps(version_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        registry_version = f"{REGISTRY_CONTRACT_VERSION}-sha256-{version_digest}"

        specs = {resource.name: resource for resource in self.resources}
        entity_records = self._read_entity_workbook(
            paths["Entity_List.xlsx"], specs["Entity_List.xlsx"], registry_version
        )
        attribute_records = self._read_csv_records(
            paths["Attributes_List.csv"], specs["Attributes_List.csv"], registry_version
        )
        watch_records = self._read_watch_records(
            paths["Disambiguation_Watch_List.csv"],
            specs["Disambiguation_Watch_List.csv"],
            registry_version,
        )
        issues = self._build_issues(entity_records, attribute_records, watch_records)
        self._snapshot = RegistrySnapshot(
            registry_version=registry_version,
            resource_hashes=actual_hashes,
            entity_records=tuple(entity_records),
            attribute_records=tuple(attribute_records),
            watch_records=tuple(watch_records),
            issues=tuple(issues),
        )
        self._build_indexes()

    def _read_entity_workbook(
        self,
        path: Path,
        spec: RegistryResourceSpec,
        registry_version: str,
    ) -> list[RegistryRecord]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        records: list[RegistryRecord] = []
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, ())
            headers = _headers(header_row)
            for row_number, row in enumerate(rows, start=2):
                values = [_cell_text(value) for value in row]
                if not any(values):
                    continue
                fields = _registry_fields(headers, values)
                record_id = _first_field_value(fields, "ID")
                headword = _first_field_value(fields, "Headword")
                records.append(
                    RegistryRecord(
                        kind="entity",
                        record_id=record_id,
                        headword=headword,
                        normalized_headword=normalize_registry_name(headword),
                        fields=fields,
                        provenance=RegistryProvenance(
                            source_file=spec.name,
                            source_path=str(spec.relative_path),
                            source_hash=spec.sha256,
                            registry_version=registry_version,
                            sheet=worksheet.title,
                            row=row_number,
                            row_end=row_number,
                        ),
                    )
                )
        workbook.close()
        return records

    def _read_csv_records(
        self,
        path: Path,
        spec: RegistryResourceSpec,
        registry_version: str,
    ) -> list[RegistryRecord]:
        records: list[RegistryRecord] = []
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            header_row = next(reader, ())
            headers = _headers(header_row)
            previous_end = reader.line_num
            for row in reader:
                start_line = previous_end + 1
                end_line = reader.line_num
                previous_end = end_line
                values = [_cell_text(value) for value in row]
                if not any(values):
                    continue
                fields = _registry_fields(headers, values)
                record_id = _first_field_value(fields, "ID")
                headword = _first_field_value(fields, "Headword")
                records.append(
                    RegistryRecord(
                        kind="attribute",
                        record_id=record_id,
                        headword=headword,
                        normalized_headword=normalize_registry_name(headword),
                        fields=fields,
                        provenance=RegistryProvenance(
                            source_file=spec.name,
                            source_path=str(spec.relative_path),
                            source_hash=spec.sha256,
                            registry_version=registry_version,
                            row=start_line,
                            row_end=end_line,
                        ),
                    )
                )
        return records

    def _read_watch_records(
        self,
        path: Path,
        spec: RegistryResourceSpec,
        registry_version: str,
    ) -> list[RegistryRecord]:
        records: list[RegistryRecord] = []
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.reader(stream)
            headers: list[str] | None = None
            previous_end = 0
            for row in reader:
                start_line = previous_end + 1
                end_line = reader.line_num
                previous_end = end_line
                values = [_cell_text(value) for value in row]
                if headers is None:
                    if values and values[0] == "Word":
                        headers = _headers(values)
                    continue
                if not any(values):
                    continue
                fields = _registry_fields(headers, values)
                record_id = _first_field_value(fields, "ID")
                headword = _first_field_value(fields, "Headword")
                records.append(
                    RegistryRecord(
                        kind="watch",
                        record_id=record_id,
                        headword=headword,
                        normalized_headword=normalize_registry_name(headword),
                        fields=fields,
                        provenance=RegistryProvenance(
                            source_file=spec.name,
                            source_path=str(spec.relative_path),
                            source_hash=spec.sha256,
                            registry_version=registry_version,
                            row=start_line,
                            row_end=end_line,
                        ),
                    )
                )
        if headers is None:
            raise RegistryIntegrityError(f"Watch List header row was not found in {path}")
        return records

    def _build_issues(
        self,
        entities: list[RegistryRecord],
        attributes: list[RegistryRecord],
        watch: list[RegistryRecord],
    ) -> list[RegistryIssue]:
        issues: list[RegistryIssue] = []
        entity_by_id = _records_by_id(entities)
        attribute_by_id = _records_by_id(attributes)
        combined = {**entity_by_id, **attribute_by_id}

        for record_id, records in sorted(entity_by_id.items()):
            if len(records) > 1:
                issues.append(
                    RegistryIssue(
                        code="duplicate_id",
                        message=f"ID {record_id} occurs in multiple Entity List rows.",
                        record_id=record_id,
                        provenance=tuple(record.provenance for record in records),
                        details={"headwords": [record.headword for record in records]},
                    )
                )
        for record in entities + attributes:
            if not record.headword:
                issues.append(
                    RegistryIssue(
                        code="empty_headword",
                        message=f"Registry row {record.record_id or '<no ID>'} has an empty Headword.",
                        record_id=record.record_id,
                        provenance=(record.provenance,),
                    )
                )

        for record in watch:
            referenced_id = record.record_id
            if not referenced_id:
                issues.append(
                    RegistryIssue(
                        code="missing_reference",
                        message="Watch List row has no referenced ID.",
                        provenance=(record.provenance,),
                    )
                )
                continue
            referenced = combined.get(referenced_id, ())
            if not referenced:
                expected_external = bool(
                    re.fullmatch(r"E-\d{4}", referenced_id) or referenced_id.casefold() == "no id"
                )
                issues.append(
                    RegistryIssue(
                        code="missing_reference",
                        message=(
                            f"Watch List ID {referenced_id} is not present in the built-in Entity or Attribute registry."
                        ),
                        record_id=referenced_id,
                        headword=record.headword,
                        provenance=(record.provenance,),
                        details={"expected_external_or_no_id": expected_external},
                    )
                )
                continue
            if len(referenced) > 1:
                issues.append(
                    RegistryIssue(
                        code="watch_reference_ambiguous_id",
                        message=f"Watch List ID {referenced_id} points to a duplicate registry ID.",
                        record_id=referenced_id,
                        headword=record.headword,
                        provenance=(record.provenance,) + tuple(item.provenance for item in referenced),
                    )
                )
            if not any(item.headword == record.headword for item in referenced):
                issues.append(
                    RegistryIssue(
                        code="watch_headword_string_difference",
                        message=(
                            f"Watch List Headword {record.headword!r} differs from the registry string(s) for {referenced_id}."
                        ),
                        record_id=referenced_id,
                        headword=record.headword,
                        provenance=(record.provenance,) + tuple(item.provenance for item in referenced),
                        details={"registry_headwords": [item.headword for item in referenced]},
                    )
                )
        return issues

    def _build_indexes(self) -> None:
        snapshot = self.snapshot
        self._id_index = _freeze_index(_records_by_id(snapshot.entity_records))
        self._attribute_id_index = _freeze_index(_records_by_id(snapshot.attribute_records))
        headwords: dict[str, list[RegistryRecord]] = defaultdict(list)
        watch: dict[str, list[RegistryRecord]] = defaultdict(list)
        terms: dict[tuple[str, str, int | None], tuple[str, ...]] = {}
        for record in snapshot.entity_records:
            if record.normalized_headword:
                headwords[record.normalized_headword].append(record)
            search_values = [record.headword]
            for field_name in ("Trigger", "Surname"):
                for value in record.values(field_name):
                    search_values.extend(_split_variants(value))
            terms[self._record_key(record)] = tuple(value for value in search_values if value)
        for record in snapshot.watch_records:
            watch_values = list(record.values("Word")) + [record.headword]
            for value in record.values("Possible Forms (non-exhaustive)"):
                watch_values.extend(_split_variants(value))
            for value in watch_values:
                normalized = normalize_registry_name(value)
                if normalized:
                    watch[normalized].append(record)
        self._headword_index = _freeze_index(headwords)
        self._watch_index = _freeze_index(watch)
        self._entity_search_terms = terms

    def _issues_for_candidates(
        self,
        candidates: tuple[RegistryRecord, ...],
        *,
        record_id: str = "",
        headword: str = "",
    ) -> tuple[RegistryIssue, ...]:
        candidate_keys = {self._record_key(record) for record in candidates}
        result = []
        for issue in self.snapshot.issues:
            issue_keys = {
                (item.source_file, item.sheet, item.row)
                for item in issue.provenance
            }
            if issue_keys & candidate_keys or (record_id and issue.record_id == record_id):
                result.append(issue)
            elif headword and normalize_registry_name(issue.headword) == normalize_registry_name(headword):
                result.append(issue)
        return tuple(_deduplicate_issues(result))

    def _lookup_result(
        self,
        query: str,
        normalized_query: str,
        status: str,
        match_type: str,
        candidates: tuple[RegistryRecord, ...],
        issues: tuple[RegistryIssue, ...],
    ) -> LookupResult:
        return LookupResult(
            query=query,
            normalized_query=normalized_query,
            status=status,
            match_type=match_type,
            candidates=candidates,
            ambiguities=issues,
            registry_version=self.registry_version,
            resource_hashes=self.resource_hashes,
        )

    @staticmethod
    def _record_key(record: RegistryRecord) -> tuple[str, str, int | None]:
        return (record.provenance.source_file, record.provenance.sheet, record.provenance.row)

    @staticmethod
    def _candidate_sort_key(candidate: CandidateMatch) -> tuple[Any, ...]:
        layer_order = {"exact": 0, "normalized": 1, "lexical_fuzzy": 2}
        return (
            layer_order.get(candidate.layer, 3),
            -candidate.score,
            candidate.record.normalized_headword,
            candidate.record.record_id,
            candidate.record.provenance.sheet,
            candidate.record.provenance.row or 0,
        )


def build_default_entity_registry_service() -> EntityRegistryService:
    return EntityRegistryService()


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _headers(values: Iterable[Any]) -> list[str]:
    headers = []
    seen: dict[str, int] = defaultdict(int)
    for index, value in enumerate(values, start=1):
        base = _cell_text(value) or f"_column_{index}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}__{seen[base]}")
    return headers


def _registry_fields(headers: list[str], values: list[str]) -> tuple[RegistryField, ...]:
    width = max(len(headers), len(values))
    fields = []
    for index in range(width):
        name = headers[index] if index < len(headers) else f"_column_{index + 1}"
        value = values[index] if index < len(values) else ""
        fields.append(RegistryField(column=index + 1, name=name, value=value))
    return tuple(fields)


def _first_field_value(fields: tuple[RegistryField, ...], name: str) -> str:
    wanted = normalize_registry_name(name)
    for field in fields:
        if normalize_registry_name(field.name) == wanted:
            return field.value
    return ""


def _records_by_id(records: Iterable[RegistryRecord]) -> dict[str, list[RegistryRecord]]:
    result: dict[str, list[RegistryRecord]] = defaultdict(list)
    for record in records:
        if record.record_id:
            result[record.record_id].append(record)
    return result


def _freeze_index(index: dict[str, list[RegistryRecord]]) -> dict[str, tuple[RegistryRecord, ...]]:
    return {key: tuple(value) for key, value in index.items()}


def _split_variants(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[;\n]", str(value or "")) if item.strip()]


def _deduplicate_issues(issues: Iterable[RegistryIssue]) -> list[RegistryIssue]:
    result = []
    seen = set()
    for issue in issues:
        key = (
            issue.code,
            issue.record_id,
            issue.headword,
            tuple((item.source_file, item.sheet, item.row) for item in issue.provenance),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(issue)
    return result
